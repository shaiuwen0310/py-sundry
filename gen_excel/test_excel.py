from openpyxl import Workbook
import datetime

datetime_dt = datetime.datetime.today()
datetime_str = datetime_dt.strftime("%Y%m%d_%H%M%S")

wb = Workbook()
ws = wb.active
ws.title = "sheet name"
ws['A1'], ws['B1'], ws['C1'] = 'number1', 'number2', 'number3'

try:
    # 可用loop去插值
    ws.append(["first", "second", "third"])
    # 動態調整row和column
    # ws.cell(row=row_num, column=(i + 1 + 3), value=qrcode_brightness_list[i])

finally:
    # 在finally儲存檔案
    wb.save(datetime_str + '.xlsx')
